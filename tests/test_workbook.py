"""Structural checks on the generated workbook."""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "tools"))

from build_protocol_assets import compute_example, compute_i4  # noqa: E402

XLSX = ROOT / "templates" / "CheckDam_Recharge_Calculator.xlsx"


def test_workbook_has_expected_sheets_and_formulas():
    assert XLSX.exists(), "run python3 tools/build_protocol_assets.py first"
    wb = load_workbook(XLSX, data_only=False)
    for name in (
        "README",
        "Site",
        "StageAreaVolume",
        "FillAndHold",
        "Evaporation",
        "DailyPond",
        "DailyCalc",
        "Wells",
        "Karez",
        "Summary",
        "Scorecard",
    ):
        assert name in wb.sheetnames, name

    assert "FillAndHold" in wb.sheetnames
    assert wb["FillAndHold"]["B6"].value == 70
    assert "SQRT" in str(wb["FillAndHold"]["B24"].value)
    assert wb["Site"]["B10"].value == 2.5
    assert wb["Site"]["B15"].value == 0.10
    assert wb["Summary"]["C5"].value == "=DailyCalc!S2"
    assert "AVERAGEIF" in str(wb["DailyCalc"]["S1"].value)
    assert "MeanMDWIR" in str(wb["DailyCalc"]["N6"].value)
    assert "MINIFS" in str(wb["Wells"]["M5"].value)
    rows, mdwir, i1, vcrest, fillings = compute_example()
    i4, _, _ = compute_i4()
    assert abs(wb["Summary"]["B24"].value - i1) < 1e-6
    assert abs(wb["Summary"]["B25"].value - mdwir) < 1e-6
    assert abs(wb["Summary"]["B26"].value - i4) < 1e-3
    assert wb["DailyPond"]["A5"].value == rows[0]["date"]
    dry_used = str(wb["DailyCalc"]["L6"].value)
    assert "C6<J6" in dry_used.replace(" ", "")
    wb.close()


if __name__ == "__main__":
    test_workbook_has_expected_sheets_and_formulas()
    print("ok test_workbook_has_expected_sheets_and_formulas")
