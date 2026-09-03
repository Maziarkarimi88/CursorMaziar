"""
Build protocol figures and the check-dam recharge workbook.

Run from the repository root:

    python3 tools/build_protocol_assets.py

Outputs:
    figures/sampling_layout.png
    figures/sampling_layout.svg
    figures/calculation_flow.png
    templates/CheckDam_Recharge_Calculator.xlsx
"""

from __future__ import annotations

from pathlib import Path

from datetime import date, timedelta

import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.patches import Arc, Circle, FancyBboxPatch, Polygon, Rectangle
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parents[1]
FIG = ROOT / "figures"
TPL = ROOT / "templates"


# ---------------------------------------------------------------------------
# Hydrology helpers (also used to seed the example year)
# ---------------------------------------------------------------------------

STAGE = [0.00, 0.50, 1.00, 1.50, 2.00, 2.50]  # m, crest = 2.50
AREA = [0, 800, 1800, 3000, 4200, 5500]  # m2
# Trapezoidal volumes
VOL = [0.0]
for i in range(1, len(STAGE)):
    dV = (STAGE[i] - STAGE[i - 1]) / 2.0 * (AREA[i] + AREA[i - 1])
    VOL.append(VOL[-1] + dV)

E_DEFAULT = {
    1: 2.5,
    2: 3.5,
    3: 5.0,
    4: 6.5,
    5: 8.0,
    6: 9.0,
    7: 9.5,
    8: 9.0,
    9: 7.0,
    10: 5.0,
    11: 3.5,
    12: 2.5,
}


def lerp(x, xs, ys):
    if x <= xs[0]:
        return ys[0]
    if x >= xs[-1]:
        return ys[-1]
    for i in range(1, len(xs)):
        if x <= xs[i]:
            t = (x - xs[i - 1]) / (xs[i] - xs[i - 1])
            return ys[i - 1] + t * (ys[i] - ys[i - 1])
    return ys[-1]


def area_of(h):
    return lerp(h, STAGE, AREA)


def vol_of(h):
    return lerp(h, STAGE, VOL)


def make_example_days():
    """Spring snowmelt filling: two rain days, one spill, then ~55 mm/day drawdown.

    Daily stage drop of 55 mm with E ≈ 5 mm/day gives MDWIR near 50 mm/day
    (about 10 × evaporation), a gravel-apex pond rather than a silted one.
    """
    days = [
        ("2025-03-08", 22.0, 1.70, "N", "N"),
        ("2025-03-09", 14.0, 2.35, "N", "N"),
        ("2025-03-10", 0.0, 2.52, "Y", "N"),
    ]
    h = 2.46
    d = date(2025, 3, 11)
    while h >= 0.18:
        iso = d.isoformat()
        if iso == "2025-03-22":
            h = min(h + 0.10, 2.50)
            days.append((iso, 8.0, round(h, 2), "N", "N"))
            h -= 0.055
        else:
            days.append((iso, 0.0, round(h, 2), "N", "N"))
            h -= 0.055
        d += timedelta(days=1)
    days.append((d.isoformat(), 0.0, 0.00, "N", "N"))
    return days


EXAMPLE_DAYS = make_example_days()

# id, class, rim elev m asl, DTW pre, DTW peak, DTW late
WELL_SERIES = [
    ("W-N1", "W-N", 1012.40, 14.20, 12.40, 13.10),
    ("W-N2", "W-N", 1011.80, 13.50, 11.85, 12.55),
    ("W-M1", "W-M", 1008.10, 16.00, 14.90, 15.40),
    ("W-M2", "W-M", 1007.40, 17.20, 16.05, 16.60),
    ("W-F1", "W-F", 1002.00, 22.40, 21.90, 22.20),
    ("C1", "C", 1010.50, 15.10, 14.70, 14.95),
    ("C2", "C", 1006.20, 18.40, 18.00, 18.25),
    ("KS", "KS", 1009.70, 11.80, 10.50, 11.10),
]


def compute_i4():
    """Seasonal rise = pre DTW − peak DTW (same as max−min water elevation)."""
    treated, control = [], []
    for _wid, cls, _rim, dtw_pre, dtw_peak, _late in WELL_SERIES:
        dh = dtw_pre - dtw_peak
        if cls in ("W-N", "W-M"):
            treated.append(dh)
        elif cls == "C":
            control.append(dh)
    t = sum(treated) / len(treated)
    c = sum(control) / len(control)
    return t - c, t, c


def compute_example():
    """Independent water-balance used to verify the spreadsheet."""
    rows = []
    dry_mdwir = []
    E_by_month = {3: 5.0, 4: 6.5}
    prev_h = None
    for date, P, h, overflow, pumped in EXAMPLE_DAYS:
        month = int(date.split("-")[1])
        E = E_by_month[month]
        A = area_of(h)
        V = vol_of(h)
        rec = {
            "date": date,
            "P": P,
            "h": h,
            "overflow": overflow,
            "pumped": pumped,
            "E_mm": E,
            "A": A,
            "V": V,
            "dry_day": (
                P == 0
                and overflow == "N"
                and pumped == "N"
                and prev_h is not None
                and h < prev_h
                and h > 0
            ),
        }
        if rec["dry_day"]:
            Aavg = 0.5 * (area_of(prev_h) + A)
            dV = vol_of(prev_h) - V
            evap_m3 = (E / 1000.0) * Aavg
            inf_m3 = dV - evap_m3
            mdwir = inf_m3 / Aavg * 1000.0 if Aavg > 0 else 0.0
            rec.update(Aavg=Aavg, dV=dV, evap_m3=evap_m3, inf_m3=inf_m3, mdwir=mdwir)
            if mdwir > 0:
                dry_mdwir.append(mdwir)
        rows.append(rec)
        prev_h = h
    mean_mdwir = sum(dry_mdwir) / len(dry_mdwir)
    i1 = 0.0
    prev_h = None
    for rec in rows:
        if prev_h is None:
            rec["inf_used"] = 0.0
            prev_h = rec["h"]
            continue
        Aavg = 0.5 * (area_of(prev_h) + rec["A"])
        if rec["dry_day"]:
            inf = rec["inf_m3"]
        else:
            inf = mean_mdwir / 1000.0 * Aavg
        rec["inf_used"] = inf
        i1 += inf
        prev_h = rec["h"]
    return rows, mean_mdwir, i1, VOL[-1], i1 / VOL[-1]


# ---------------------------------------------------------------------------
# Figures
# ---------------------------------------------------------------------------

def _style_axes(ax):
    ax.set_aspect("equal")
    ax.axis("off")


def build_sampling_layout():
    FIG.mkdir(exist_ok=True)
    fig, ax = plt.subplots(figsize=(11.5, 8.2), dpi=160)
    ax.set_xlim(0, 16)
    ax.set_ylim(0, 11.2)
    _style_axes(ax)
    fig.patch.set_facecolor("white")

    # Mountains
    peaks = Polygon(
        [(0.3, 9.2), (1.4, 10.7), (2.6, 9.4), (3.7, 10.8), (5.0, 9.3),
         (6.2, 10.6), (7.4, 9.2), (7.4, 11.15), (0.3, 11.15)],
        closed=True, facecolor="#c5cdd6", edgecolor="#5c6770", lw=1.0,
    )
    ax.add_patch(peaks)
    ax.text(3.8, 10.85, "Foothills / catchment", ha="center", va="center",
            fontsize=9, color="#333", fontstyle="italic")

    # Treated fan
    fan = Polygon(
        [(2.2, 8.55), (0.4, 1.3), (7.6, 1.3), (5.6, 8.55)],
        closed=True, facecolor="#f3e6c8", edgecolor="#b0894a", lw=1.2, alpha=0.95,
    )
    ax.add_patch(fan)
    ax.text(4.0, 2.15, "Treated alluvial fan\n(coarse gravel → sand down-fan)",
            ha="center", fontsize=8, color="#6b4f1d")

    # Control fan
    cfan = Polygon(
        [(9.6, 8.55), (8.3, 1.3), (15.4, 1.3), (13.0, 8.55)],
        closed=True, facecolor="#e8e8e8", edgecolor="#7a7a7a", lw=1.2, ls="--",
    )
    ax.add_patch(cfan)
    ax.text(11.8, 2.15, "Control fan (no check dam)\n≥ 2 km laterally",
            ha="center", fontsize=8, color="#444")

    # Wadis
    ax.annotate("", xy=(3.9, 1.45), xytext=(3.9, 8.7),
                arrowprops=dict(arrowstyle="-", color="#3a7ca5", lw=3.5))
    ax.plot([3.9, 3.9], [8.7, 9.3], color="#3a7ca5", lw=3.5)
    ax.annotate("", xy=(11.5, 1.45), xytext=(11.5, 8.7),
                arrowprops=dict(arrowstyle="-", color="#8aa4b5", lw=2.5, ls="--"))

    ax.annotate("Flow / groundwater\ndown-fan", xy=(4.55, 4.7), fontsize=7,
                color="#3a7ca5", rotation=90, va="center")
    ax.text(4.25, 8.85, "Ephemeral wadi", fontsize=7, color="#215973")
    ax.text(11.75, 8.85, "Untreated wadi", fontsize=7, color="#555")

    # Pond + dam
    pond = FancyBboxPatch((3.15, 7.55), 1.5, 0.85, boxstyle="round,pad=0.02,rounding_size=0.2",
                          facecolor="#7ec8e3", edgecolor="#1d5673", lw=1.2)
    ax.add_patch(pond)
    ax.add_patch(Rectangle((3.05, 7.42), 1.7, 0.16, facecolor="#5c4033", edgecolor="black", lw=0.6))
    ax.text(3.9, 7.95, "Pond", ha="center", va="center", fontsize=8, color="#123")
    ax.text(3.9, 7.50, "Check dam  (SG)", ha="center", va="bottom", fontsize=7.5, color="white",
            fontweight="bold")

    # Distance arcs
    for r, label in [(0.55, "200 m"), (1.55, "800 m"), (2.55, "1.5 km")]:
        ax.add_patch(Arc((3.9, 7.50), 2 * r, 2 * r * 1.15, theta1=200, theta2=340,
                         edgecolor="#9a7b4f", lw=0.7, ls=":"))
        ax.text(3.9 + r + 0.15, 7.35 - r * 0.55, label, fontsize=6.5, color="#9a7b4f")

    def well(x, y, code, color, fill):
        ax.add_patch(Circle((x, y), 0.13, facecolor=fill, edgecolor=color, lw=1.2, zorder=5))
        ax.text(x + 0.22, y, code, fontsize=7.5, va="center", color=color, fontweight="bold", zorder=5)

    # Treated wells
    well(3.15, 7.15, "W-N1", "#1b4332", "#d8f3dc")
    well(4.65, 7.15, "W-N2", "#1b4332", "#d8f3dc")
    well(3.35, 6.15, "W-M1", "#1b4332", "#95d5b2")
    well(4.55, 5.95, "W-M2", "#1b4332", "#95d5b2")
    well(3.55, 4.55, "W-F1", "#1b4332", "#52b788")
    well(4.45, 3.55, "W-F2", "#1b4332", "#52b788")

    # Karez
    ax.plot([3.55, 5.55], [6.55, 3.25], color="#d9480f", lw=2.2, zorder=4)
    ax.plot([3.55, 5.55], [6.55, 3.25], color="#d9480f", lw=2.2, ls="None", marker="o",
            markersize=3.5, markevery=0.2, zorder=4)
    ax.add_patch(Circle((3.55, 6.55), 0.16, facecolor="#fff3bf", edgecolor="#d9480f", lw=1.4, zorder=6))
    ax.text(2.05, 6.55, "KS  sarchah\n(mother well)", fontsize=7, color="#d9480f", va="center")
    ax.add_patch(Circle((5.55, 3.25), 0.16, facecolor="#fff3bf", edgecolor="#d9480f", lw=1.4, zorder=6))
    ax.text(5.85, 3.25, "KO  owkura\n(karez outlet)", fontsize=7, color="#d9480f", va="center")
    ax.text(5.7, 5.15, "Karez gallery\n~1–2 km", fontsize=7, color="#d9480f", rotation=-52)

    # Village + fields
    ax.add_patch(FancyBboxPatch((5.9, 2.7), 1.35, 0.85, boxstyle="round,pad=0.02",
                                facecolor="#ffd6a5", edgecolor="#9a5b13", lw=0.8))
    ax.text(6.57, 3.12, "Village\n+ RG", ha="center", va="center", fontsize=7.5)
    ax.add_patch(Rectangle((5.85, 1.55), 1.55, 0.9, facecolor="#b7e4c7", edgecolor="#2d6a4f", lw=0.6))
    ax.text(6.62, 2.0, "Fields", ha="center", va="center", fontsize=7.5, color="#1b4332")

    # Control wells
    well(10.85, 6.35, "C1", "#495057", "#dee2e6")
    well(12.15, 5.15, "C2", "#495057", "#dee2e6")
    well(11.15, 3.85, "C3", "#495057", "#dee2e6")
    ax.text(12.6, 6.5, "Control wells\n(same fan position)", fontsize=7.5, color="#495057")

    # Rain / BM callouts
    ax.annotate("SG staff gauge\non spillway wall", xy=(4.65, 7.55), xytext=(6.3, 8.35),
                fontsize=7.5, color="#1d5673",
                arrowprops=dict(arrowstyle="->", color="#1d5673", lw=0.8))
    ax.annotate("RG rain gauge\nin open compound", xy=(6.9, 3.4), xytext=(7.7, 4.35),
                fontsize=7.5, color="#1d5673",
                arrowprops=dict(arrowstyle="->", color="#1d5673", lw=0.8))

    # Title and legend
    ax.text(8.0, 11.0, "Well and karez sampling layout — Kandahar / Zabul alluvial fan",
            ha="center", fontsize=12, fontweight="bold")
    ax.text(8.0, 10.55, "Not to scale. Place real GPS on Form A. Typical influence 0.2–1.5 km down-fan.",
            ha="center", fontsize=8, color="#444")

    handles = [
        mpatches.Patch(facecolor="#7ec8e3", edgecolor="#1d5673", label="Check-dam pond"),
        mpatches.Patch(facecolor="#d8f3dc", edgecolor="#1b4332", label="Treated wells (W-N, W-M, W-F)"),
        mpatches.Patch(facecolor="#fff3bf", edgecolor="#d9480f", label="Karez (KS, KO)"),
        mpatches.Patch(facecolor="#dee2e6", edgecolor="#495057", label="Control wells (C)"),
        mpatches.Patch(facecolor="#f3e6c8", edgecolor="#b0894a", label="Treated fan"),
        mpatches.Patch(facecolor="#e8e8e8", edgecolor="#7a7a7a", label="Control fan"),
    ]
    ax.legend(handles=handles, loc="lower left", bbox_to_anchor=(0.01, 0.01),
              frameon=True, fontsize=7.5, ncol=3, fancybox=False, edgecolor="#ccc")

    ax.text(15.7, 10.35, "HILLS", fontsize=8, color="#5c6770", ha="right")
    ax.text(15.7, 1.45, "REGISTAN /\nLOWER FAN", fontsize=7, color="#7a7a7a", ha="right", va="bottom")

    fig.tight_layout(pad=0.4)
    png = FIG / "sampling_layout.png"
    svg = FIG / "sampling_layout.svg"
    fig.savefig(png, bbox_inches="tight", facecolor="white")
    fig.savefig(svg, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    return png


def build_calculation_flow():
    fig, ax = plt.subplots(figsize=(11.2, 4.6), dpi=150)
    ax.set_xlim(0, 12)
    ax.set_ylim(0, 4.4)
    ax.axis("off")
    fig.patch.set_facecolor("white")

    boxes = [
        (0.2, 2.4, "Form B\nstage, rain\n07:00 photo"),
        (2.5, 2.4, "Stage–area–\nvolume table\n(Form A)"),
        (4.8, 2.4, "Dry-day\nMDWIR\n(I3)"),
        (7.1, 2.4, "Apply MDWIR\nto all ponded\ndays → I1, I2"),
        (9.4, 2.4, "Scorecard\nI1–I5\n(Form D)"),
        (2.5, 0.45, "Form C\nwells Δh\nvs control → I4"),
        (7.1, 0.45, "Karez KO\nflow days\nand Q → I5"),
    ]
    for x, y, t in boxes:
        ax.add_patch(FancyBboxPatch((x, y), 2.05, 1.55, boxstyle="round,pad=0.04,rounding_size=0.08",
                                    facecolor="#eef6fb", edgecolor="#1d5673", lw=1.1))
        ax.text(x + 1.02, y + 0.78, t, ha="center", va="center", fontsize=8)

    arrows = [
        ((2.25, 3.15), (2.5, 3.15)),
        ((4.55, 3.15), (4.8, 3.15)),
        ((6.85, 3.15), (7.1, 3.15)),
        ((9.15, 3.15), (9.4, 3.15)),
        ((3.52, 2.4), (3.52, 2.0)),
        ((8.12, 2.4), (8.12, 2.0)),
        ((4.55, 1.2), (9.4, 2.4)),
        ((9.15, 1.2), (10.4, 2.4)),
    ]
    for (x1, y1), (x2, y2) in arrows:
        ax.annotate("", xy=(x2, y2), xytext=(x1, y1),
                    arrowprops=dict(arrowstyle="->", color="#1d5673", lw=1.1))

    ax.set_title("Calculation flow (workbook sheets DailyPond → DailyCalc → Summary → Scorecard)",
                 fontsize=10, pad=6)
    fig.tight_layout()
    png = FIG / "calculation_flow.png"
    fig.savefig(png, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    return png


# ---------------------------------------------------------------------------
# Workbook
# ---------------------------------------------------------------------------

THIN = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)
NAVY = "1D5673"
SAND = "F3E6C8"
GREEN = "D8F3DC"
BLUE = "D6EAF8"
AMBER = "FFF3BF"
GREY = "F4F4F4"
INPUT_FILL = PatternFill("solid", fgColor="FFF8E7")
CALC_FILL = PatternFill("solid", fgColor="EAF4FB")
HEAD_FILL = PatternFill("solid", fgColor=NAVY)
HEAD_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
TITLE_FONT = Font(name="Calibri", size=16, bold=True, color=NAVY)
SECTION_FONT = Font(name="Calibri", size=12, bold=True, color=NAVY)
LABEL_FONT = Font(name="Calibri", size=10)
NUM_FONT = Font(name="Calibri", size=10)


def _header_row(ws, row, headers, fill=HEAD_FILL):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row, i, h)
        c.fill = fill
        c.font = HEAD_FONT
        c.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
        c.border = THIN
    ws.row_dimensions[row].height = 28


def _input(cell, value=None, num_fmt=None):
    if value is not None:
        cell.value = value
    cell.fill = INPUT_FILL
    cell.border = THIN
    cell.font = NUM_FONT
    if num_fmt:
        cell.number_format = num_fmt


def _calc(cell, value=None, num_fmt=None):
    if value is not None:
        cell.value = value
    cell.fill = CALC_FILL
    cell.border = THIN
    cell.font = NUM_FONT
    if num_fmt:
        cell.number_format = num_fmt


def _label(cell, text):
    cell.value = text
    cell.font = LABEL_FONT
    cell.alignment = Alignment(vertical="center")


def build_workbook():
    TPL.mkdir(exist_ok=True)
    rows, mean_mdwir, i1, vcrest, fillings = compute_example()
    i4_py, _, _ = compute_i4()

    wb = Workbook()

    # ----- README -----
    ws = wb.active
    ws.title = "README"
    ws.sheet_properties.tabColor = NAVY
    ws["A1"] = "Check-dam groundwater recharge calculator"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:G1")
    ws["A2"] = (
        "Field workbook for rainfall-fed check dams designed to hold about 1–3 months when full. "
        "Yellow cells are inputs. Blue cells are formulas — do not type over them. "
        "An example spring filling is pre-loaded so you can see a complete result; "
        "replace it with your Form B and Form C data. Sheet FillAndHold estimates rainfall to fill 100%."
    )
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A2:G4")
    ws["A6"] = "Sheet order"
    ws["A6"].font = SECTION_FONT
    steps = [
        ("1. Site", "Names, crest, catchment ha, Sy. Matches Form A."),
        ("2. StageAreaVolume", "h–A–V table from the empty-pond survey. Do not skip."),
        ("3. FillAndHold", "Rainfall (mm) to fill 100%; days to empty from full (1–3 month check)."),
        ("4. Evaporation", "Monthly open-water E (mm/day). Override if you have a pan (use 0.7 × pan)."),
        ("5. DailyPond", "Paste Form B: date, rain, stage, overflow, pumped."),
        ("6. DailyCalc", "Interpolates A and V, flags dry days, computes MDWIR and daily infiltration."),
        ("7. Wells", "Form C water levels. Converts DTW to elevation and seasonal Δh."),
        ("8. Karez", "Weekly KO discharge and flow-day count."),
        ("9. Summary", "I1–I5 for the year."),
        ("10. Scorecard", "Interpretation and desilting decision."),
    ]
    _header_row(ws, 7, ["Sheet", "Purpose"])
    for i, (a, b) in enumerate(steps, 8):
        ws.cell(i, 1, a).border = THIN
        ws.cell(i, 2, b).border = THIN
        ws.cell(i, 2).alignment = Alignment(wrap_text=True)
    ws["A18"] = "Colour key"
    ws["A18"].font = SECTION_FONT
    ws["A19"] = "Input"
    ws["A19"].fill = INPUT_FILL
    ws["B19"] = "Type here (field data)"
    ws["A20"] = "Calculated"
    ws["A20"].fill = CALC_FILL
    ws["B20"] = "Formula — leave it"
    ws["A22"] = "Dry-day rule (Dashora / MARVI)"
    ws["A22"].font = SECTION_FONT
    ws["A23"] = (
        "A day is DRY if rain = 0, overflow = N, pumped = N, and stage fell. "
        "Infiltration volume = drop in storage − evaporation × mean area. "
        "MDWIR = that volume / mean area. "
        "On wet or spilling days, infiltration = mean dry-day MDWIR × mean area. "
        "I1 is the sum of daily infiltration over all days with water in the pond."
    )
    ws["A23"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A23:G25")
    ws["A27"] = "Example currently loaded"
    ws["A27"].font = SECTION_FONT
    ws["A28"] = (
        f"Example spring filling (Mar–Apr), crest volume {vcrest:,.0f} m³. "
        f"Independent Python check: I1 = {i1:,.0f} m³, I2 = {fillings:.2f} fillings, "
        f"mean MDWIR = {mean_mdwir:.1f} mm/day. "
        "After you build this file, open Summary and confirm it matches within 1%."
    )
    ws["A28"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A28:G30")
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 88
    for col in "CDEFG":
        ws.column_dimensions[col].width = 14
    ws.row_dimensions[2].height = 48
    ws.freeze_panes = "A8"

    # ----- Site -----
    ws = wb.create_sheet("Site")
    ws.sheet_properties.tabColor = "B0894A"
    ws["A1"] = "Site (Form A)"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")
    ws["A3"] = "Item"
    ws["B3"] = "Value"
    ws["A3"].font = HEAD_FONT
    ws["B3"].font = HEAD_FONT
    ws["A3"].fill = HEAD_FILL
    ws["B3"].fill = HEAD_FILL
    labels = [
        "Check dam ID", "Province / district", "Wadi / fan name", "Treated or control",
        "Year built", "Cascade position (1 = upstream)", "Crest stage on gauge (m)",
        "Gauge zero", "Latitude", "Longitude", "Catchment area (ha)",
        "Sy default (fan gravel)", "Sy low", "Sy high", "Observer name", "Technician name",
    ]
    values = [
        "CD-EX-01", "Kandahar / Arghandab example", "Example fan apex wadi", "Treated",
        2024, 1, 2.50, "Pond bed", 31.62, 65.74, 450,
        0.10, 0.08, 0.15, "Example observer", "Example technician",
    ]
    fmts = [None, None, None, None, "0", "0", "0.00", None, "0.000", "0.000", "0.0",
            "0.00", "0.00", "0.00", None, None]
    for i, (lab, val, fmt) in enumerate(zip(labels, values, fmts), 4):
        _label(ws.cell(i, 1), lab)
        _input(ws.cell(i, 2), val, fmt)
    ws["A21"] = "Other sheets read B10 = crest stage (m), B14 = catchment (ha), B15 = Sy default."
    ws["A21"].font = Font(italic=True, size=9, color="666666")
    # named range via wb - we'll just use Site!B10 etc.
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 34
    ws["D3"] = "Notes"
    ws["D3"].font = SECTION_FONT
    ws["D4"] = (
        "Replace the example ID before a real campaign. "
        "Sy = 0.10 is the Afghan karez-model default for fan gravels. "
        "Always report I4 as a range using Sy low and Sy high."
    )
    ws["D4"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("D4:F8")

    # ----- StageAreaVolume -----
    ws = wb.create_sheet("StageAreaVolume")
    ws.sheet_properties.tabColor = "3A7CA5"
    ws["A1"] = "Stage–area–volume (Form A4)"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = (
        "Enter surveyed stage (m above gauge zero), surface area (m²), then volume is computed "
        "by the trapezoid rule. Keep stages in increasing order. Crest must appear in the table."
    )
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A2:E3")
    _header_row(ws, 5, ["Stage h (m)", "Area A (m²)", "ΔV (m³)", "Cumulative V (m³)", "Notes"])
    for i, (h, a, v) in enumerate(zip(STAGE, AREA, VOL)):
        r = 6 + i
        _input(ws.cell(r, 1), h, "0.00")
        _input(ws.cell(r, 2), a, "0")
        if i == 0:
            _calc(ws.cell(r, 3), 0, "0.0")
            _calc(ws.cell(r, 4), 0, "0.0")
        else:
            # ΔV = (h_i - h_{i-1})/2 * (A_i + A_{i-1})
            _calc(ws.cell(r, 3), f"=IF(A{r}=\"\",\"\",(A{r}-A{r-1})/2*(B{r}+B{r-1}))", "0.0")
            _calc(ws.cell(r, 4), f"=IF(A{r}=\"\",\"\",D{r-1}+C{r})", "0.0")
        ws.cell(r, 5).border = THIN
        if abs(h - 2.5) < 1e-9:
            ws.cell(r, 5).value = "CREST"
    # extra empty rows for longer surveys
    for i in range(6):
        r = 12 + i
        _input(ws.cell(r, 1), None, "0.00")
        _input(ws.cell(r, 2), None, "0")
        _calc(ws.cell(r, 3), f"=IF(OR(A{r}=\"\",A{r-1}=\"\"),\"\",(A{r}-A{r-1})/2*(B{r}+B{r-1}))", "0.0")
        _calc(ws.cell(r, 4), f"=IF(A{r}=\"\",\"\",IF(D{r-1}=\"\",C{r},D{r-1}+C{r}))", "0.0")
        ws.cell(r, 5).border = THIN
    ws["A20"] = "Crest volume (m³)"
    ws["B20"] = '=INDEX(D6:D17,MATCH(Site!B10,A6:A17,0))'
    _calc(ws["B20"], ws["B20"].value, "0.0")
    ws["C20"] = "Looks up V at Site crest stage. If #N/A, put an exact crest row in the table."
    ws["A21"] = "Crest area (m²)"
    ws["B21"] = '=INDEX(B6:B17,MATCH(Site!B10,A6:A17,0))'
    _calc(ws["B21"], ws["B21"].value, "0")
    for col, w in zip("ABCDE", [16, 16, 16, 20, 28]):
        ws.column_dimensions[col].width = w

    chart = LineChart()
    chart.title = "Area and volume versus stage"
    chart.style = 10
    chart.y_axis.title = "A (m²) or V (m³)"
    chart.x_axis.title = "Stage (m)"
    chart.height = 8
    chart.width = 15
    data = Reference(ws, min_col=2, min_row=5, max_col=4, max_row=11)
    cats = Reference(ws, min_col=1, min_row=6, max_row=11)
    chart.add_data(data, from_rows=False, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    ws.add_chart(chart, "G5")

    # ----- Evaporation -----
    ws = wb.create_sheet("Evaporation")
    ws.sheet_properties.tabColor = "E07A3D"
    ws["A1"] = "Open-water evaporation (mm/day)"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = (
        "Defaults are for Kandahar open water when no pan is available. "
        "If you have a Class A pan, enter 0.70 × monthly mean pan in column B."
    )
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A2:D3")
    _header_row(ws, 5, ["Month number", "E (mm/day)", "Month name", "Source"])
    names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    for m in range(1, 13):
        r = 5 + m
        _input(ws.cell(r, 1), m, "0")
        _input(ws.cell(r, 2), E_DEFAULT[m], "0.0")
        ws.cell(r, 3, names[m - 1]).border = THIN
        ws.cell(r, 4, "Default Kandahar open water").border = THIN
    ws["A20"] = "Lookup: DailyCalc uses =VLOOKUP(month, Evaporation!A6:B17, 2, FALSE)"
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 32

    # ----- DailyPond -----
    ws = wb.create_sheet("DailyPond")
    ws.sheet_properties.tabColor = "2D6A4F"
    ws["A1"] = "Daily pond log (Form B) — INPUT"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = (
        "One row per day. Leave stage blank if not visited. Write 0 for a dry pond (empty). "
        "Overflow and Pumped must be Y or N. Example filling is rows 5–34; overwrite these."
    )
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A2:H3")
    headers = ["Date", "Rain P (mm)", "Stage h (m)", "Overflow (Y/N)", "Pumped (Y/N)",
               "Photo ID", "Notes", "Observer"]
    _header_row(ws, 4, headers)
    dv = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
    dv.error = "Enter Y or N"
    dv.errorTitle = "Overflow / Pumped"
    ws.add_data_validation(dv)
    dv.add("D5:E204")
    n_example = len(EXAMPLE_DAYS)
    for i, (date, P, h, ov, pu) in enumerate(EXAMPLE_DAYS):
        r = 5 + i
        _input(ws.cell(r, 1), date)
        ws.cell(r, 1).number_format = "YYYY-MM-DD"
        _input(ws.cell(r, 2), P, "0.0")
        _input(ws.cell(r, 3), h, "0.00")
        _input(ws.cell(r, 4), ov)
        _input(ws.cell(r, 5), pu)
        _input(ws.cell(r, 6), f"IMG-{i+1:03d}")
        ws.cell(r, 7).fill = INPUT_FILL
        ws.cell(r, 7).border = THIN
        _input(ws.cell(r, 8), "Example")
    for r in range(5 + n_example, 205):
        for c in range(1, 9):
            ws.cell(r, c).fill = INPUT_FILL
            ws.cell(r, c).border = THIN
        ws.cell(r, 1).number_format = "YYYY-MM-DD"
        ws.cell(r, 2).number_format = "0.0"
        ws.cell(r, 3).number_format = "0.00"
    widths = [14, 14, 14, 16, 14, 14, 28, 14]
    for col, w in zip("ABCDEFGH", widths):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = "A4:H204"

    # ----- DailyCalc -----
    # Linear interpolation of A and V from StageAreaVolume using MATCH/INDEX.
    # Table is A6:D17 on StageAreaVolume.
    ws = wb.create_sheet("DailyCalc")
    ws.sheet_properties.tabColor = "1D5673"
    ws["A1"] = "Daily water balance — CALCULATED (do not type here)"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = (
        "h_prev is previous row stage. DryDay = 1 only if rain=0, overflow=N, pumped=N, and stage fell. "
        "MDWIR_mm_d is computed on dry days. InfiltrationUsed_m3 uses the dry-day value on dry days "
        "and MeanMDWIR × Aavg on other ponded days (Dashora shortcut)."
    )
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A2:P3")
    calc_headers = [
        "Date", "Rain_mm", "h_m", "Overflow", "Pumped",
        "Month", "E_mm", "A_m2", "V_m3", "h_prev", "Aavg_m2",
        "DryDay", "MDWIR_mm_d", "InfiltrationUsed_m3", "Evap_m3", "Flag",
    ]
    _header_row(ws, 4, calc_headers)
    # Helper interpolation:
    # MATCH(h, stages, 1) = index of largest stage <= h
    # We use a named approach inline.
    # For h=0, MATCH may still work if 0 is first.
    sav_h = "StageAreaVolume!$A$6:$A$17"
    sav_a = "StageAreaVolume!$B$6:$B$17"
    sav_v = "StageAreaVolume!$D$6:$D$17"

    def interp_formula(hcell, yrange):
        # linear interpolation; if h blank return blank
        # i = MATCH(h, stages, 1)
        # x0 = INDEX(stages,i); x1 = INDEX(stages,i+1)
        # y0 = INDEX(y,i); y1 = INDEX(y,i+1)
        return (
            f'IF({hcell}="","",'
            f'IF({hcell}<=INDEX({sav_h},1),INDEX({yrange},1),'
            f'IF({hcell}>=INDEX({sav_h},COUNTA({sav_h})),INDEX({yrange},COUNTA({sav_h})),'
            f'LET(i,MATCH({hcell},{sav_h},1),'
            f'x0,INDEX({sav_h},i),x1,INDEX({sav_h},i+1),'
            f'y0,INDEX({yrange},i),y1,INDEX({yrange},i+1),'
            f'y0+({hcell}-x0)/(x1-x0)*(y1-y0)))))'
        )

    # Excel LET is available in Excel 365 / 2021. For broader compatibility, avoid LET.
    # Use a simpler FORECAST-style with MATCH without LET.
    def interp_compat(hcell, yrange):
        return (
            f'IF({hcell}="","",'
            f'IF({hcell}<=INDEX({sav_h},1),INDEX({yrange},1),'
            f'IF({hcell}>=INDEX({sav_h},COUNTA({sav_h})),INDEX({yrange},COUNTA({sav_h})),'
            f'(INDEX({yrange},MATCH({hcell},{sav_h},1))+'
            f'({hcell}-INDEX({sav_h},MATCH({hcell},{sav_h},1)))/'
            f'(INDEX({sav_h},MATCH({hcell},{sav_h},1)+1)-INDEX({sav_h},MATCH({hcell},{sav_h},1)))*'
            f'(INDEX({yrange},MATCH({hcell},{sav_h},1)+1)-INDEX({yrange},MATCH({hcell},{sav_h},1))))))))'
        )

    for r in range(5, 205):
        src = r  # DailyPond row
        # Date, rain, h, overflow, pumped
        _calc(ws.cell(r, 1), f'=IF(DailyPond!A{src}="","",DailyPond!A{src})')
        ws.cell(r, 1).number_format = "YYYY-MM-DD"
        _calc(ws.cell(r, 2), f'=IF(A{r}="","",DailyPond!B{src})', "0.0")
        _calc(ws.cell(r, 3), f'=IF(A{r}="","",DailyPond!C{src})', "0.00")
        _calc(ws.cell(r, 4), f'=IF(A{r}="","",DailyPond!D{src})')
        _calc(ws.cell(r, 5), f'=IF(A{r}="","",DailyPond!E{src})')
        _calc(ws.cell(r, 6), f'=IF(A{r}="","",MONTH(A{r}))', "0")
        _calc(ws.cell(r, 7), f'=IF(A{r}="","",VLOOKUP(F{r},Evaporation!$A$6:$B$17,2,FALSE))', "0.0")
        _calc(ws.cell(r, 8), f'=IF(A{r}="","",{interp_compat(f"C{r}", sav_a)})', "0")
        _calc(ws.cell(r, 9), f'=IF(A{r}="","",{interp_compat(f"C{r}", sav_v)})', "0.0")
        if r == 5:
            _calc(ws.cell(r, 10), f'=IF(A{r}="","",NA())')  # no previous
            _calc(ws.cell(r, 11), f'=IF(A{r}="","",H{r})', "0")
            _calc(ws.cell(r, 12), 0, "0")
            _calc(ws.cell(r, 13), '=""')
            _calc(ws.cell(r, 14), 0, "0.0")
            _calc(ws.cell(r, 15), 0, "0.0")
            _calc(ws.cell(r, 16), '="first"')
        else:
            _calc(ws.cell(r, 10), f'=IF(A{r}="","",IF(A{r-1}="",NA(),C{r-1}))', "0.00")
            _calc(ws.cell(r, 11), f'=IF(OR(A{r}="",ISNA(J{r})),IF(A{r}="","",H{r}),(H{r-1}+H{r})/2)', "0")
            # DryDay
            _calc(
                ws.cell(r, 12),
                f'=IF(OR(A{r}="",ISNA(J{r})),0,IF(AND(B{r}=0,D{r}="N",E{r}="N",C{r}<J{r},C{r}>0),1,0))',
                "0",
            )
            # MDWIR on dry days: ((Vprev-V) - E/1000*Aavg) / Aavg * 1000
            _calc(
                ws.cell(r, 13),
                f'=IF(L{r}<>1,"",IF(K{r}<=0,"",(I{r-1}-I{r}-G{r}/1000*K{r})/K{r}*1000))',
                "0.00",
            )
            # Evap m3
            _calc(ws.cell(r, 15), f'=IF(A{r}="","",G{r}/1000*K{r})', "0.0")
            # InfiltrationUsed: if dry use computed; elseif ponded (h>0 or hprev>0) use mean MDWIR
            # Mean MDWIR is $S$1 computed below... we'll put MeanMDWIR in S1 after the loop.
            _calc(
                ws.cell(r, 14),
                f'=IF(A{r}="","",IF(L{r}=1,I{r-1}-I{r}-O{r},'
                f'IF(OR(C{r}>0,N(J{r})>0),MeanMDWIR/1000*K{r},0)))',
                "0.0",
            )
            _calc(
                ws.cell(r, 16),
                f'=IF(A{r}="","",IF(L{r}=1,"dry",IF(D{r}="Y","spill",IF(B{r}>0,"wet",IF(E{r}="Y","pumped","other")))))',
            )

    # Mean MDWIR named cell
    ws["R1"] = "MeanMDWIR_mm_d"
    ws["S1"] = "=IFERROR(AVERAGEIF(L5:L204,1,M5:M204),0)"
    _calc(ws["S1"], ws["S1"].value, "0.00")
    ws["R2"] = "I1 sum infiltration m3"
    ws["S2"] = "=SUM(N5:N204)"
    _calc(ws["S2"], ws["S2"].value, "0.0")
    ws["R3"] = "Days ponded"
    ws["S3"] = '=COUNTIF(C5:C204,">0")'
    _calc(ws["S3"], ws["S3"].value, "0")
    ws["R4"] = "Dry days used"
    ws["S4"] = "=SUM(L5:L204)"
    _calc(ws["S4"], ws["S4"].value, "0")
    wb.defined_names.add(DefinedName(name="MeanMDWIR", attr_text="DailyCalc!$S$1"))

    ws["R6"] = "Python check (example)"
    ws["S6"] = mean_mdwir
    ws["S6"].number_format = "0.00"
    ws["T6"] = "Expected mean MDWIR mm/d"
    ws["R7"] = "Python I1 m3"
    ws["S7"] = i1
    ws["S7"].number_format = "0.0"
    ws["R8"] = "Difference I1 % "
    ws["S8"] = '=IF(S7=0,"", (S2-S7)/S7)'
    ws["S8"].number_format = "0.0%"
    _calc(ws["S8"], ws["S8"].value, "0.0%")

    for i, w in enumerate([12, 10, 10, 11, 10, 9, 9, 11, 11, 10, 11, 10, 13, 18, 11, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.column_dimensions["R"].width = 24
    ws.column_dimensions["S"].width = 14
    ws.column_dimensions["T"].width = 26
    ws.freeze_panes = "A5"

    # Highlight dry days
    green_fill = PatternFill("solid", fgColor="D8F3DC")
    ws.conditional_formatting.add(
        "L5:L204",
        FormulaRule(formula=["L5=1"], fill=green_fill),
    )

    # ----- Wells -----
    ws = wb.create_sheet("Wells")
    ws.sheet_properties.tabColor = "52B788"
    ws["A1"] = "Wells (Form C) — INPUT + Δh"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = (
        "One row per visit per well. WaterElevation = RimElev_masl − DTW. "
        "Seasonal rise for each well is MAX−MIN elevation in this sheet. "
        "I4 uses mean rise of W-N and W-M minus mean rise of C."
    )
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A2:J3")
    _header_row(ws, 4, [
        "Date", "Well ID", "Class (W-N/W-M/W-F/C/KS)", "DTW (m)",
        "Rim elev (m asl)", "Water elev (m asl)", "Pump on (Y/N)", "EC (µS/cm)", "Notes",
    ])
    # Example well time series: pre-flood 1 Mar, peak 25 Mar, recession 20 Apr
    # Treated near: rise 1.8 m; mid 1.1; far 0.5; control 0.4; sarchah 1.3
    dates = ["2025-03-01", "2025-03-25", "2025-04-20"]
    r = 5
    for wid, cls, rim, d1, d2, d3 in WELL_SERIES:
        for date, dtw in zip(dates, (d1, d2, d3)):
            _input(ws.cell(r, 1), date)
            ws.cell(r, 1).number_format = "YYYY-MM-DD"
            _input(ws.cell(r, 2), wid)
            _input(ws.cell(r, 3), cls)
            _input(ws.cell(r, 4), dtw, "0.00")
            _input(ws.cell(r, 5), rim, "0.00")
            _calc(ws.cell(r, 6), f'=IF(D{r}="","",E{r}-D{r})', "0.00")
            _input(ws.cell(r, 7), "N")
            _input(ws.cell(r, 8), None, "0")
            ws.cell(r, 9).fill = INPUT_FILL
            ws.cell(r, 9).border = THIN
            r += 1
    for rr in range(r, 205):
        for c in range(1, 10):
            ws.cell(rr, c).fill = INPUT_FILL
            ws.cell(rr, c).border = THIN
        _calc(ws.cell(rr, 6), f'=IF(D{rr}="","",E{rr}-D{rr})', "0.00")
        ws.cell(rr, 1).number_format = "YYYY-MM-DD"
        ws.cell(rr, 4).number_format = "0.00"
        ws.cell(rr, 5).number_format = "0.00"

    # Pivot-like summary by well using AVERAGEIF/MAXIFS/MINIFS
    ws["K4"] = "Well ID"
    ws["L4"] = "Class"
    ws["M4"] = "Min elev"
    ws["N4"] = "Max elev"
    ws["O4"] = "Seasonal Δh (m)"
    for col in "KLMNO":
        ws[f"{col}4"].fill = HEAD_FILL
        ws[f"{col}4"].font = HEAD_FONT
        ws[f"{col}4"].border = THIN
    for i, (wid, cls, *_) in enumerate(WELL_SERIES):
        rr = 5 + i
        _input(ws.cell(rr, 11), wid)
        _calc(ws.cell(rr, 12), f'=IFERROR(INDEX($C$5:$C$204,MATCH(K{rr},$B$5:$B$204,0)),"")')
        _calc(ws.cell(rr, 13), f'=IF(K{rr}="","",MINIFS($F$5:$F$204,$B$5:$B$204,K{rr}))', "0.00")
        _calc(ws.cell(rr, 14), f'=IF(K{rr}="","",MAXIFS($F$5:$F$204,$B$5:$B$204,K{rr}))', "0.00")
        _calc(ws.cell(rr, 15), f'=IF(K{rr}="","",N{rr}-M{rr})', "0.00")

    ws["K15"] = "Mean Δh W-N and W-M (m)"
    ws["L15"] = (
        '=(SUMIF(L5:L12,"W-N",O5:O12)+SUMIF(L5:L12,"W-M",O5:O12))'
        '/(COUNTIF(L5:L12,"W-N")+COUNTIF(L5:L12,"W-M"))'
    )
    _calc(ws["L15"], ws["L15"].value, "0.00")
    ws["K16"] = "Mean Δh control C (m)"
    ws["L16"] = '=IFERROR(AVERAGEIF(L5:L12,"C",O5:O12),0)'
    _calc(ws["L16"], ws["L16"].value, "0.00")
    ws["K17"] = "I4 extra rise (m)"
    ws["L17"] = "=L15-L16"
    _calc(ws["L17"], ws["L17"].value, "0.00")
    ws["K18"] = "Recharge depth Sy default (mm)"
    ws["L18"] = "=L17*Site!B15*1000"
    _calc(ws["L18"], ws["L18"].value, "0.0")
    ws["K19"] = "Recharge depth Sy low (mm)"
    ws["L19"] = "=L17*Site!B16*1000"
    _calc(ws["L19"], ws["L19"].value, "0.0")
    ws["K20"] = "Recharge depth Sy high (mm)"
    ws["L20"] = "=L17*Site!B17*1000"
    _calc(ws["L20"], ws["L20"].value, "0.0")

    i4_py, _, _ = compute_i4()
    ws["K22"] = "Python check I4 (m)"
    ws["L22"] = round(i4_py, 3)
    ws["L22"].number_format = "0.000"

    for col, w in zip("ABCDEFGHIJKLMNO", [12, 12, 22, 12, 14, 16, 14, 12, 20, 12, 32, 14, 12, 12, 16]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A5"

    # ----- Karez -----
    ws = wb.create_sheet("Karez")
    ws.sheet_properties.tabColor = "D9480F"
    ws["A1"] = "Karez outlet (Form C) — INPUT"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Weekly (or daily while ponded) discharge at owkura. Q = 0 means dry. Do not enter the tunnel."
    ws.merge_cells("A2:F3")
    _header_row(ws, 4, ["Date", "Karez name", "Q (L/s)", "Flowing (1/0)", "Control Q (L/s)", "Notes"])
    karez_q = [
        ("2025-03-01", 0.0),
        ("2025-03-08", 0.4),
        ("2025-03-15", 1.8),
        ("2025-03-22", 2.4),
        ("2025-03-29", 2.1),
        ("2025-04-05", 1.5),
        ("2025-04-12", 0.9),
        ("2025-04-19", 0.3),
        ("2025-04-26", 0.0),
        ("2025-05-03", 0.0),
    ]
    control_q = [0, 0, 0.2, 0.3, 0.2, 0.1, 0, 0, 0, 0]
    for i, ((date, q), cq) in enumerate(zip(karez_q, control_q)):
        r = 5 + i
        _input(ws.cell(r, 1), date)
        ws.cell(r, 1).number_format = "YYYY-MM-DD"
        _input(ws.cell(r, 2), "Example karez")
        _input(ws.cell(r, 3), q, "0.00")
        _calc(ws.cell(r, 4), f'=IF(C{r}="","",IF(C{r}>0,1,0))', "0")
        _input(ws.cell(r, 5), cq, "0.00")
        ws.cell(r, 6).fill = INPUT_FILL
        ws.cell(r, 6).border = THIN
    for r in range(15, 105):
        for c in range(1, 7):
            ws.cell(r, c).fill = INPUT_FILL
            ws.cell(r, c).border = THIN
        _calc(ws.cell(r, 4), f'=IF(C{r}="","",IF(C{r}>0,1,0))', "0")
        ws.cell(r, 1).number_format = "YYYY-MM-DD"
        ws.cell(r, 3).number_format = "0.00"
        ws.cell(r, 5).number_format = "0.00"

    ws["H4"] = "Treated flow days (visits)"
    ws["I4"] = "=SUM(D5:D104)"
    _calc(ws["I4"], ws["I4"].value, "0")
    ws["H5"] = "Mean Q while flowing (L/s)"
    ws["I5"] = '=IFERROR(AVERAGEIF(D5:D104,1,C5:C104),0)'
    _calc(ws["I5"], ws["I5"].value, "0.00")
    ws["H6"] = "Control flow days"
    ws["I6"] = '=COUNTIF(E5:E104,">0")'
    _calc(ws["I6"], ws["I6"].value, "0")
    ws["H7"] = "I5 extra flow-day visits"
    ws["I7"] = "=I4-I6"
    _calc(ws["I7"], ws["I7"].value, "0")
    ws["H8"] = "Note"
    ws["I8"] = "If visits are weekly, multiply extra visits × 7 for extra days (Summary does this)."
    ws.merge_cells("I8:L8")
    for col, w in zip("ABCDEFGHI", [12, 18, 12, 14, 16, 24, 12, 32, 14]):
        ws.column_dimensions[col].width = w

    # ----- FillAndHold (rainfall to 100% full; days of storage) -----
    ws = wb.create_sheet("FillAndHold")
    ws.sheet_properties.tabColor = "2A9D8F"
    ws["A1"] = "Rainfall to fill 100% and days water will stay"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:E1")
    ws["A2"] = (
        "Design hold is about 1–3 months when full. Yellow = assumptions. "
        "Blue uses Site catchment (ha) and crest V, A from StageAreaVolume. "
        "See docs/STORAGE_DURATION_AND_FILLING.md. SCS-CN: USDA-NRCS NEH-630."
    )
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A2:E3")

    ws["A5"] = "Inputs (yellow)"
    ws["A5"].font = SECTION_FONT
    fill_inputs = [
        (6, "Curve number CN", 70, "0.0", "70–85 rocky rangeland; 60–75 fair pasture"),
        (7, "Runoff coefficient C", 0.15, "0.00", "First look only; 0.08–0.30 typical dryland"),
        (8, "Initial abstraction ratio λ", 0.20, "0.00", "0.20 standard SCS; try 0.05 in arid lands"),
        (9, "Extra loss while filling (fraction)", 0.15, "0.00", "Infiltration + E during the fill storm"),
        (10, "Volume already in pond (m³)", 0, "0.0", "0 if starting empty"),
        (11, "Open-water E (mm/day)", 6.0, "0.0", "Pan × 0.7; winter lower, summer higher"),
        (12, "Bed infiltration i (mm/day)", 25.0, "0.0", "Use MDWIR once you have Form B"),
        (13, "Wall/outlet leak (m³/day)", 0, "0.0", "0 if masonry and drain closed"),
    ]
    for r, lab, val, fmt, note in fill_inputs:
        _label(ws.cell(r, 1), lab)
        _input(ws.cell(r, 2), val, fmt)
        ws.cell(r, 3, note).alignment = Alignment(wrap_text=True)

    ws["A15"] = "Results (blue)"
    ws["A15"].font = SECTION_FONT
    results = [
        (16, "Crest storage V (m³)", "=StageAreaVolume!B20", "0.0"),
        (17, "Crest water area A (m²)", "=StageAreaVolume!B21", "0"),
        (18, "Catchment area (m²)", "=Site!B14*10000", "0"),
        (19, "Volume still needed (m³)", "=(B16-B10)*(1+B9)", "0.0"),
        (20, "Required runoff depth Q (mm)", "=IF(B18=0,\"\",B19/B18*1000)", "0.00"),
        (21, "P to fill, constant C (mm)", "=IF(OR(B7=0,B20=\"\"),\"\",B20/B7)", "0.0"),
        (22, "SCS S (mm)", "=IF(OR(B6<=0,B6>=100),\"\",25400/B6-254)", "0.0"),
        (23, "SCS Ia (mm)", "=IF(B22=\"\",\"\",B8*B22)", "0.0"),
        (24, "P to fill, SCS-CN (mm)", "=IF(OR(B20=\"\",B22=\"\"),\"\",IF(B20<=0,B23,B23+(B20+SQRT(B20^2+4*B20*B22))/2))", "0.0"),
        (25, "Capacity as mm on catchment", "=IF(B18=0,\"\",B16/B18*1000)", "0.00"),
        (26, "Days to empty, full area (shortest)", "=IF((B11+B12)/1000*B17+B13<=0,\"\",B16/((B11+B12)/1000*B17+B13))", "0.0"),
        (27, "Days to empty, mean area (mid)", "=IF((B11+B12)/1000*(B17/2)+B13<=0,\"\",B16/((B11+B12)/1000*(B17/2)+B13))", "0.0"),
    ]
    for r, lab, formula, fmt in results:
        _label(ws.cell(r, 1), lab)
        _calc(ws.cell(r, 2), formula, fmt)

    ws["A29"] = (
        "P to fill is ONE storm from a dry catchment. Many 5–10 mm showers may never fill "
        "the dam if each is below Ia. Days to empty assume no further inflow and no pumping. "
        "Full-area days is a low estimate; mean-area is closer to a shrinking pond. "
        "Compare B26–B27 with the 30–90 day design. If B26 is << 30, the bed is too leaky "
        "or the pond is too shallow. If I/E is near 1, you have an evaporation pan."
    )
    ws["A29"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A29:E32")
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 52
    ws.row_dimensions[2].height = 36
    ws.row_dimensions[29].height = 48

    # ----- Summary -----
    ws = wb.create_sheet("Summary")
    ws.sheet_properties.tabColor = NAVY
    ws["A1"] = "Annual summary — I1 to I5"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = '=CONCATENATE("Dam ",Site!B4," — ",Site!B5," — ",Site!B6)'
    ws["A2"].font = SECTION_FONT
    ws.merge_cells("A2:D2")

    _header_row(ws, 4, ["ID", "Indicator", "Value", "Unit", "How it is computed"])
    summary_rows = [
        ("I1", "Infiltration volume", "=DailyCalc!S2", "m³", "Sum of daily infiltration (dry-day storage drop − E, else MDWIR × A)"),
        ("I2", "Fillings per year", "=IF(StageAreaVolume!B20=0,\"\",C5/StageAreaVolume!B20)", "—", "I1 / crest storage"),
        ("I3", "Mean MDWIR", "=DailyCalc!S1", "mm/day", "Average dry-day infiltration rate"),
        ("", "Mean E while ponded", "=IFERROR(AVERAGEIF(DailyCalc!C5:C204,\">0\",DailyCalc!G5:G204),0)", "mm/day", "Evaporation on days with water"),
        ("", "MDWIR / E", "=IF(C8=0,\"\",C7/C8)", "—", "Healthy gravel beds are often 4–8"),
        ("I4", "Extra water-table rise vs control", "=Wells!L17", "m", "Mean Δh of W-N and W-M minus mean Δh of C"),
        ("", "Recharge depth (Sy default)", "=Wells!L18", "mm", "Sy × I4; not a volume"),
        ("", "Recharge depth range", '=TEXT(Wells!L19,"0.0")&" – "&TEXT(Wells!L20,"0.0")', "mm", "Sy low to Sy high"),
        ("I5", "Extra karez flow-day visits", "=Karez!I7", "visits", "Treated flowing visits − control flowing visits"),
        ("", "Extra karez days if weekly", "=C13*7", "days", "Assumes weekly KO visits"),
        ("", "Mean KO discharge while flowing", "=Karez!I5", "L/s", "Average of Q > 0"),
        ("", "Days with water in pond", "=DailyCalc!S3", "days", "Stage > 0"),
        ("", "Dry days used for MDWIR", "=DailyCalc!S4", "days", "Should be as many as possible"),
        ("", "Crest storage", "=StageAreaVolume!B20", "m³", "From survey"),
        ("", "Catchment rain in DailyPond", "=SUM(DailyPond!B5:B204)", "mm", "Village gauge total for logged days"),
        ("", "I1 as depth on catchment", "=IF(Site!B14=0,\"\",C5/(Site!B14*10))", "mm", "m³ / (ha × 10) = mm"),
    ]
    # Wait Site!B14 is catchment area - yes row 14.
    for i, (sid, name, formula, unit, how) in enumerate(summary_rows):
        r = 5 + i
        ws.cell(r, 1, sid).border = THIN
        ws.cell(r, 1).font = Font(bold=True, name="Calibri")
        ws.cell(r, 2, name).border = THIN
        _calc(ws.cell(r, 3), formula)
        if "range" in name:
            ws.cell(r, 3).number_format = "@"
        elif unit in ("m³", "mm", "mm/day"):
            ws.cell(r, 3).number_format = "0.0"
        elif unit in ("m", "L/s"):
            ws.cell(r, 3).number_format = "0.00"
        elif unit in ("days", "visits", "—"):
            ws.cell(r, 3).number_format = "0.00"
        ws.cell(r, 4, unit).border = THIN
        ws.cell(r, 5, how).border = THIN
        ws.cell(r, 5).alignment = Alignment(wrap_text=True)
        if sid:
            ws.cell(r, 1).fill = PatternFill("solid", fgColor=AMBER)

    ws["A23"] = "Workbook vs independent Python check (example dataset only)"
    ws["A23"].font = SECTION_FONT
    ws["A24"] = "Python I1 (m³)"
    _calc(ws["B24"], i1, "0.0")
    ws["C24"] = "Workbook I1 − Python"
    _calc(ws["D24"], "=C5-B24", "0.0")
    ws["A25"] = "Python mean MDWIR (mm/d)"
    _calc(ws["B25"], mean_mdwir, "0.00")
    ws["C25"] = "Workbook MDWIR − Python"
    _calc(ws["D25"], "=C7-B25", "0.00")
    ws["A26"] = "Python I4 (m)"
    _calc(ws["B26"], i4_py, "0.000")
    ws["C26"] = "Workbook I4 − Python"
    _calc(ws["D26"], "=C10-B26", "0.000")
    ws["A27"] = (
        "If the example is still loaded, differences should be near zero "
        "(interpolation rounding < ~2%). After you replace the example, ignore this block."
    )
    ws["A27"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A27:E28")

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 62
    for r in range(5, 21):
        ws.row_dimensions[r].height = 22

    # ----- Scorecard -----
    ws = wb.create_sheet("Scorecard")
    ws.sheet_properties.tabColor = "B0894A"
    ws["A1"] = "Annual scorecard (copy onto Form D)"
    ws["A1"].font = TITLE_FONT
    ws["A3"] = "Test"
    ws["B3"] = "Result"
    ws["C3"] = "Suggested decision"
    for col in "ABC":
        ws[f"{col}3"].fill = HEAD_FILL
        ws[f"{col}3"].font = HEAD_FONT
        ws[f"{col}3"].border = THIN
        ws[f"{col}3"].alignment = Alignment(wrap_text=True)
    ws["A4"] = "Is MDWIR at least 4 × evaporation?"
    ws["B4"] = '=IF(Summary!C9>=4,"YES — bed is infiltrating","NO — pond may be evaporating or silted")'
    ws["C4"] = "If NO: schedule manual desilting this dry season. Do not compact with machinery."
    ws["A5"] = "Did treated wells rise > 0.3 m more than control?"
    ws["B5"] = '=IF(Summary!C10>0.3,"YES — local water table responded","NO — response not detected")'
    ws["C5"] = "If NO but I1 is large: check new pumps, wait one more year for unsaturated-zone lag, inspect karez."
    ws["A6"] = "Did the karez flow more than the control?"
    ws["B6"] = '=IF(Summary!C13>0,"YES — extra flow visits","NO — no extra karez flow")'
    ws["C6"] = "If wells rose but karez did not: gallery may be above the water table or collapsed."
    ws["A7"] = "Fillings per year"
    ws["B7"] = "=Summary!C6"
    ws["B7"].number_format = "0.00"
    ws["C7"] = "Gujarat/Rajasthan analogues often 1–7. Below 1 in a normal year is poor siting or a clogged bed."
    ws["A8"] = "Decision (edit by hand after the committee meeting)"
    _input(ws["B8"], "Maintain only — example dataset looks functional")
    ws.merge_cells("B8:C8")
    for r in range(4, 9):
        for c in range(1, 4):
            ws.cell(r, c).border = THIN
            ws.cell(r, c).alignment = Alignment(wrap_text=True, vertical="center")
        ws.row_dimensions[r].height = 36
        if r < 8:
            _calc(ws.cell(r, 2), ws.cell(r, 2).value)
            ws.cell(r, 3).fill = PatternFill("solid", fgColor=GREY)
    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 62

    ws["A10"] = "Do not claim regional groundwater recovery from one dam. Report I1–I5 and the map of which wells rose."
    ws["A10"].font = Font(italic=True, size=10, color="666666")
    ws.merge_cells("A10:C10")

    # print settings
    for sheet in wb.worksheets:
        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.fitToPage = True
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 1
        sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.print_title_rows = "1:4" if sheet.title in (
            "DailyPond", "DailyCalc", "Wells", "Karez"
        ) else None

    out = TPL / "CheckDam_Recharge_Calculator.xlsx"
    wb.save(out)
    return out, mean_mdwir, i1, fillings, vcrest


def main():
    png1 = build_sampling_layout()
    png2 = build_calculation_flow()
    xlsx, mdwir, i1, fillings, vcrest = build_workbook()
    rows, _, _, _, _ = compute_example()
    print("Wrote", png1)
    print("Wrote", png2)
    print("Wrote", xlsx)
    print(f"Crest volume {vcrest:.1f} m3")
    print(f"Example mean MDWIR {mdwir:.2f} mm/d")
    print(f"Example I1 {i1:.1f} m3")
    print(f"Example I2 {fillings:.2f}")
    print(f"Dry days {sum(1 for r in rows if r.get('dry_day'))}")


if __name__ == "__main__":
    main()
